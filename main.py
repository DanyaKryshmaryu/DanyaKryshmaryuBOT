import telebot
import requests
import json
import xlrd
import xlwt
import openpyxl
import requests
from telebot import types
bot = telebot.TeleBot('1731471550:AAFBXeBnA4xkFmvSfXu8nUPdsLHY6UxP90s')
@bot.message_handler(commands=['start'])
def start_message(message):
    bot.send_message(message.chat.id, 'Привет')
    # открываем файл
rb = xlrd.open_workbook('..\ProjectBOT\debet.xlsx',formatting_info=True)
#выбираем активный лист
sheet = rb.sheet_by_index(0)
#получаем значение первой ячейки A1
val = sheet.row_values(0)[0]
#получаем список значений из всех записей
vals = [sheet.row_values(rownum) for rownum in range(sheet.nrows)]
wb = xlwt.Workbook()
ws = wb.add_sheet('Test')

sheet = wb['test']
#в A1 записываем значение из ячейки A1 прошлого файла
ws.write(0, 0, val[0])

#в столбец B запишем нашу последовательность из столбца A исходного файла
i = 0
for rec in vals:
    ws.write(i,1,rec[0])
    i =+ i

#сохраняем рабочую книгу
wb.save('..\ProjectBOT\debet.xlsx')
wb = openpyxl.load_workbook(filename = '..\ProjectBOT\debet.xlsx')
#считываем значение определенной ячейки
val = sheet['A1'].value

#считываем заданный диапазон
vals = [v[0].value for v in sheet.range('A1:A2')]
#записываем значение в определенную ячейку
sheet['B1'] = val

#записываем последовательность
i = 0
for rec in vals:
    sheet.cell(row=i, column=2).value = rec
    i =+ 1

# сохраняем данные
wb.save('../ArticleScripts/ExcelPython/openpyxl.xlsx')

if __name__ == "__main__":

    bot.polling(none_stop=True)


